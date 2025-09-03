
from __future__ import annotations

import base64
import io
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from PIL import Image

# ---------- Setup ----------
load_dotenv()
api_key = os.getenv("AZURE_OPENAI_API_KEY")
endpoint = "https://openai-mavericks.openai.azure.com/openai/v1/"
deployment_name = "gpt-4"
client = OpenAI(base_url=endpoint, api_key=api_key)

st.set_page_config(page_title="Wesco Journal Totals Validator", layout="wide")


# ---------- Agentic Framework ----------
@dataclass
class Step:
    name: str
    run: Callable[[Dict[str, Any]], Dict[str, Any]]


class Agent:
    def __init__(self, steps: List[Step]):
        self.steps = steps

    def execute(self, state: Dict[str, Any]) -> Dict[str, Any]:
        state.setdefault("errors", [])
        for step in self.steps:
            try:
                updates = step.run(state)
                if updates:
                    state.update(updates)
            except Exception as exc:
                # Why: keep UI flow; surface under a debug expander
                state["errors"].append(f"{step.name}: {exc}")
        return state


# ---------- Steps ----------

def prepare_image_state(state: Dict[str, Any]) -> Dict[str, Any]:
    image_file = state.get("image_file")
    if image_file is None:
        return {}
    image_bytes = image_file.read()
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")
    ext = Path(image_file.name).suffix.lower()
    if ext in (".jpg", ".jpeg"):
        image_mime = "image/jpeg"
    elif ext == ".png":
        image_mime = "image/png"
    else:
        image_mime = image_file.type or "image/png"
    return {"image_bytes": image_bytes, "image_b64": image_b64, "image_mime": image_mime}


def vision_extract_totals(state: Dict[str, Any]) -> Dict[str, Any]:
    image_b64 = state.get("image_b64")
    image_mime = state.get("image_mime", "image/png")
    if not image_b64:
        return {"image_debit": None, "image_credit": None}

    completion = client.chat.completions.create(
        model=deployment_name,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": (
                    "From this journal table image, extract ONLY the final TOTAL Debit and final TOTAL Credit "
                    "values from the summary row at the bottom of the table. Do not extract line items. "
                    "Return just the two numbers."
                )},
                {"type": "image_url", "image_url": {"url": f"data:{image_mime};base64,{image_b64}"}},
            ],
        }],
        max_tokens=200,
    )

    vision_response = completion.choices[0].message.content or ""
    m = re.findall(r"\$?([\d,]+(?:\.\d{2}))", vision_response)
    if len(m) >= 2:
        image_debit = float(m[0].replace(",", ""))
        image_credit = float(m[1].replace(",", ""))
        return {"image_debit": image_debit, "image_credit": image_credit}
    else:
        return {"image_debit": None, "image_credit": None, "vision_error": True}


# ---------- Currency helpers ----------

def normalize_currency(raw: Optional[str]) -> Optional[str]:
    """Normalize raw currency to ISO code when unambiguous; avoid guessing for '$'/'¥'."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    s = re.sub(r"^[\[\(\{\s]+|[\]\)\}\s]+$", "", s)

    m = re.search(r"\b([A-Z]{3})\b", s)
    if m:
        return m.group(1)

    name_map = {
        "US DOLLAR": "USD", "USDOLLAR": "USD",
        "EURO": "EUR", "POUND": "GBP", "POUND STERLING": "GBP",
        "INDIAN RUPEE": "INR", "RUPEE": "INR",
        "YEN": "JPY", "RENMINBI": "CNY", "RMB": "CNY", "YUAN": "CNY",
    }
    if s in name_map:
        return name_map[s]

    sym_map = {
        "€": "EUR", "£": "GBP", "₹": "INR",
        "C$": "CAD", "A$": "AUD", "NZ$": "NZD", "HK$": "HKD", "S$": "SGD", "NT$": "TWD",
    }
    for k, v in sym_map.items():
        if k in s:
            return v

    if "$" in s or "¥" in s:
        return None
    return None


def currency_symbol(norm: Optional[str], raw: Optional[str]) -> str:
    """Best-effort symbol from normalized code or raw; defaults to '$'."""
    if norm:
        m = {
            "USD": "$", "EUR": "€", "GBP": "£", "INR": "₹", "JPY": "¥", "CNY": "¥",
            "CAD": "C$", "AUD": "A$", "NZD": "NZ$", "HKD": "HK$", "SGD": "S$", "TWD": "NT$",
        }
        sym = m.get(norm)
        if sym:
            return sym
    if raw:
        for t in ["C$", "A$", "NZ$", "HK$", "S$", "NT$"]:
            if t in raw:
                return t
        for ch in ["€", "£", "₹", "¥", "$"]:
            if ch in raw:
                return ch
    return "$"


def vision_detect_currency(state: Dict[str, Any]) -> Dict[str, Any]:
    image_b64 = state.get("image_b64")
    image_mime = state.get("image_mime", "image/png")
    if not image_b64:
        return {"image_currency_norm": None, "image_currency_raw": None}

    resp = client.chat.completions.create(
        model=deployment_name,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": (
                    "From this journal table image, what currency are the totals denominated in? "
                    "Look for a currency indicator near the totals or in parentheses/brackets in headers. "
                    "Return ONLY the currency code if shown (e.g., USD, EUR, INR, GBP). "
                    "If no code is visible, return ONLY the currency symbol (e.g., $, €, £, ₹, ¥). "
                    "Return just that single token."
                )},
                {"type": "image_url", "image_url": {"url": f"data:{image_mime};base64,{image_b64}"}},
            ],
        }],
        max_tokens=10,
    )

    text = (resp.choices[0].message.content or "").strip()
    norm = normalize_currency(text)
    return {"image_currency_raw": text if text else None, "image_currency_norm": norm}


# ---------- Excel helpers ----------

def load_full_excel_as_df(filepath) -> pd.DataFrame:
    ext = Path(filepath.name).suffix.lower()
    if ext == ".xlsx":
        wb = load_workbook(filepath, data_only=True)
        sheet = wb.active
        data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        return pd.DataFrame(data)
    elif ext == ".xls":
        df_temp = pd.read_excel(filepath, engine="xlrd")
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_temp.to_excel(writer, index=False)
        buffer.seek(0)
        wb = load_workbook(buffer, data_only=True)
        sheet = wb.active
        data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        return pd.DataFrame(data)
    else:
        raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")


def find_journal_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    for idx, row in df_raw.iterrows():
        row_strs = [str(c).strip().lower() for c in row if pd.notna(c)]
        if any("entered debit" in c for c in row_strs) and any("entered credit" in c for c in row_strs):
            return idx
    return None


def extract_journal_table_step(state: Dict[str, Any]) -> Dict[str, Any]:
    excel_file = state.get("excel_file")
    if excel_file is None:
        return {"df": pd.DataFrame()}

    df_raw = load_full_excel_as_df(excel_file)
    header_row = find_journal_header_row(df_raw)
    if header_row is None:
        return {"df": pd.DataFrame()}

    df_journal = df_raw.iloc[header_row:].copy()
    df_journal.columns = df_journal.iloc[0]
    df_journal = df_journal.drop(header_row)
    df_journal = df_journal.reset_index(drop=True)
    df_journal.columns = [str(col).strip() for col in df_journal.columns]
    return {"df": df_journal}


def detect_excel_currency_step(state: Dict[str, Any]) -> Dict[str, Any]:
    df = state.get("df")
    if df is None or df.empty:
        return {"excel_currency_norm": None, "excel_currency_raw": None}

    cur_cols = [c for c in df.columns if "currency" in str(c).lower()]
    if not cur_cols:
        return {"excel_currency_norm": None, "excel_currency_raw": None}

    s = df[cur_cols[0]].astype(str).str.strip()
    s_non_empty = s[s.str.len() > 0]
    if s_non_empty.empty:
        return {"excel_currency_norm": None, "excel_currency_raw": None}

    vc = s_non_empty.value_counts()
    excel_raw = vc.index[0]
    excel_norm = normalize_currency(excel_raw)
    return {"excel_currency_raw": excel_raw, "excel_currency_norm": excel_norm}


def compute_excel_totals_step(state: Dict[str, Any]) -> Dict[str, Any]:
    df = state.get("df")
    if df is None or df.empty:
        return {"excel_debit": None, "excel_credit": None, "debit_col": None, "credit_col": None}

    debit_col = [col for col in df.columns if "entered debit" in str(col).lower()]
    credit_col = [col for col in df.columns if "entered credit" in str(col).lower()]

    if not debit_col or not credit_col:
        return {"excel_debit": None, "excel_credit": None, "debit_col": None, "credit_col": None}

    debit_col = debit_col[0]
    credit_col = credit_col[0]

    for col in [debit_col, credit_col]:
        df[col] = pd.to_numeric(
            df[col].astype(str).str.replace(",", "").str.replace("$", ""), errors="coerce"
        )

    not_both = ~(df[debit_col].notna() & df[credit_col].notna())
    row_text = df.astype(str).apply(lambda r: " | ".join(r.values).lower(), axis=1)
    not_summary = ~row_text.str.contains(r"\\btotal\\b|\\bsummary\\b", regex=True)

    context_cols = [
        col for col in df.columns if any(k in str(col).lower() for k in [
            "legal entity", "operating group", "account", "department", "site",
            "intercompany", "project", "future 1", "future 2", "currency"
        ])
    ]

    required_non_nulls = max(1, len(context_cols) // 2)
    not_mostly_empty = df[context_cols].notna().sum(axis=1) >= required_non_nulls if context_cols else True

    filtered = df[not_both & not_summary & not_mostly_empty]
    excel_debit = filtered[debit_col].sum(skipna=True)
    excel_credit = filtered[credit_col].sum(skipna=True)

    return {
        "excel_debit": float(excel_debit) if pd.notna(excel_debit) else None,
        "excel_credit": float(excel_credit) if pd.notna(excel_credit) else None,
        "debit_col": debit_col,
        "credit_col": credit_col,
    }


# ---------- UI ----------
header_left, header_right = st.columns([4, 1])
with header_left:
    st.markdown(
        """
        <h1 style='margin-bottom: 0;'>📊 Wesco Journal Totals Validator</h1>
        <p style='margin-top: 0;'>Upload a journal table image and Excel file. We'll extract the <b>TOTAL Debit</b> and <b>TOTAL Credit</b> from both, and compare them for consistency.</p>
    """,
        unsafe_allow_html=True,
    )
with header_right:
    st.image("WESCO_International_logo.png", width=100)

col1, col2 = st.columns(2)
with col1:
    st.subheader("📷 Upload Journal Image")
    image_file = st.file_uploader("PNG or JPEG", type=["png", "jpg", "jpeg"], key="image")
with col2:
    st.subheader("📄 Upload Journal Excel")
    excel_file = st.file_uploader("XLSX or XLS", type=["xlsx", "xls"], key="excel")

left, center, right = st.columns([1, 4, 1])
with center:
    run = st.button("🔍 Extract and Compare", use_container_width=True)
    st.markdown(
        """
        <style>
        div.stButton > button {
            background-color: #00A03E;
            color: white;
            font-size: 1.2em;
            padding: 0.8em;
            border-radius: 8px;
        }
        </style>
    """,
        unsafe_allow_html=True,
    )

if run:
    if not image_file or not excel_file:
        st.warning("Please upload both image and Excel files.")
        st.stop()

    steps = [
        Step(name="prepare_image_state", run=prepare_image_state),
        Step(name="vision_extract_totals", run=vision_extract_totals),
        Step(name="vision_detect_currency", run=vision_detect_currency),
        Step(name="extract_journal_table", run=extract_journal_table_step),
        Step(name="detect_excel_currency", run=detect_excel_currency_step),
        Step(name="compute_excel_totals", run=compute_excel_totals_step),
    ]
    agent = Agent(steps)

    state: Dict[str, Any] = {"image_file": image_file, "excel_file": excel_file}
    state = agent.execute(state)

    image_debit = state.get("image_debit")
    image_credit = state.get("image_credit")
    excel_debit = state.get("excel_debit")
    excel_credit = state.get("excel_credit")
    debit_col = state.get("debit_col")
    credit_col = state.get("credit_col")

    if state.get("vision_error"):
        st.error("❌ Could not extract debit/credit from image response.")

    st.markdown("---")
    st.markdown(
        """
    <div style='text-align: center;'>
        <h2 style='color: #333;'>📌 Comparison Result</h2>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # Currency awareness panel
    img_curr_norm = state.get("image_currency_norm")
    img_curr_raw = state.get("image_currency_raw")
    xls_curr_norm = state.get("excel_currency_norm")
    xls_curr_raw = state.get("excel_currency_raw")

    def _disp(v_norm, v_raw):
        return v_norm or (v_raw if v_raw else "Unknown")

    img_disp = _disp(img_curr_norm, img_curr_raw)
    xls_disp = _disp(xls_curr_norm, xls_curr_raw)

    same_currency = (
        img_curr_norm is not None and xls_curr_norm is not None and img_curr_norm == xls_curr_norm
    )
    if same_currency:
        currency_status_html = "<span style='color: green;'>✅ Currency match detected.</span>"
    elif img_disp == "Unknown" or xls_disp == "Unknown":
        currency_status_html = "<span style='color: #555;'>ℹ️ Currency could not be fully determined from one or both sources.</span>"
    else:
        currency_status_html = "<span style='color: red;'>⚠️ Currency mismatch detected. Please confirm.</span>"

    from html import escape as _esc
    st.markdown(
        (
            "<div style='margin: 0.5em 0 1em 0; padding: 0.75em; border: 1px solid #eee; border-radius: 8px;'>"
            "<strong>Detected Currency Awareness</strong><br/>"
            "<div style='display: flex; gap: 2rem; margin-top: 0.5em;'>"
            f"<div>🖼️ Image currency: <code>{_esc(str(img_disp))}</code></div>"
            f"<div>📄 Excel currency: <code>{_esc(str(xls_disp))}</code></div>"
            "</div>"
            f"<div style='margin-top: 0.5em;'>{currency_status_html}</div>"
            "</div>"
        ),
        unsafe_allow_html=True,
    )

    # Symbols for metrics
    image_symbol = currency_symbol(img_curr_norm, img_curr_raw)
    excel_symbol = currency_symbol(xls_curr_norm, xls_curr_raw)

    if None not in (image_debit, image_credit, excel_debit, excel_credit):
        col1, col2 = st.columns(2)
        with col1:
            st.metric(label="🔎 Image Total Debit", value=f"{image_symbol}{image_debit:,.2f}")
            st.metric(label="🔎 Image Total Credit", value=f"{image_symbol}{image_credit:,.2f}")
        with col2:
            st.metric(label="📊 Excel Total Debit", value=f"{excel_symbol}{excel_debit:,.2f}")
            st.metric(label="📊 Excel Total Credit", value=f"{excel_symbol}{excel_credit:,.2f}")

        match = abs(image_debit - excel_debit) < 0.01 and abs(image_credit - excel_credit) < 0.01
        if match:
            st.markdown(
                """
            <div style='text-align: center; padding: 1em;'>
                <span style='font-size: 1.5em; color: green;'>✅ MATCH: Image and Excel totals are the same.</span>
            </div>
            """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                """
            <div style='text-align: center; padding: 1em;'>
                <span style='font-size: 1.5em; color: red;'>❌ NO MATCH: The extracted totals do not align.</span>
            </div>
            """,
                unsafe_allow_html=True,
            )
    else:
        st.warning("⚠️ Unable to determine or compare totals from image and/or Excel.")

    # Preview
    st.markdown("---")
    st.markdown("---")
    st.markdown("---")
    st.markdown("### 🔍 For Reference: Preview Extracted Data")

    prev1, prev2 = st.columns(2)
    with prev1:
        st.markdown("🖼️ **Uploaded Image Preview**", unsafe_allow_html=True)
        img_bytes = state.get("image_bytes")
        if img_bytes:
            img = Image.open(io.BytesIO(img_bytes))
            fixed_height = 400
            w_percent = (fixed_height / float(img.size[1]))
            width = int((float(img.size[0]) * float(w_percent)))
            img = img.resize((width, fixed_height))
            st.image(img, use_container_width=True)
        else:
            st.info("No image preview available.")

    with prev2:
        st.markdown("📊 **Uploaded Excel Preview**", unsafe_allow_html=True)
        df = state.get("df", pd.DataFrame())
        if state.get("debit_col") and state.get("credit_col") and not df.empty:
            st.dataframe(df[[state["debit_col"], state["credit_col"]]].head(10), use_container_width=True, height=400)
        else:
            st.info("No preview available for Excel debit/credit columns.")

    if state.get("errors"):
        with st.expander("Debug: Agent step errors"):
            for e in state["errors"]:
                st.write(e)

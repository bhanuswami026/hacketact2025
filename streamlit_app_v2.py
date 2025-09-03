import streamlit as st
import pandas as pd
import base64
import os
import io
import re
from openai import OpenAI
from openpyxl import load_workbook
from dotenv import load_dotenv
from pathlib import Path
from PIL import Image

# ---------- Setup ----------
load_dotenv()
api_key = os.getenv("AZURE_OPENAI_API_KEY")
endpoint = "https://openai-mavericks.openai.azure.com/openai/v1/"
deployment_name = "gpt-4"

client = OpenAI(base_url=endpoint, api_key=api_key)

st.set_page_config(page_title="Wesco Journal Totals Validator", layout="wide")

# ---------- Wesco Branding Header ----------
header_left, header_right = st.columns([4, 1])
with header_left:
    st.markdown("""
        <h1 style='margin-bottom: 0;'>📊 Wesco Journal Totals Validator</h1>
        <p style='margin-top: 0;'>Upload a journal table image and Excel file. We'll extract the <b>TOTAL Debit</b> and <b>TOTAL Credit</b> from both, and compare them for consistency.</p>
    """, unsafe_allow_html=True)
with header_right:
    st.image("WESCO_International_logo.png", width=100)

# ---------- Upload Section ----------
col1, col2 = st.columns(2)
with col1:
    st.subheader("📷 Upload Journal Image")
    image_file = st.file_uploader("PNG only", type=["png"], key="image")
with col2:
    st.subheader("📄 Upload Journal Excel")
    excel_file = st.file_uploader("XLSX or XLS", type=["xlsx", "xls"], key="excel")

# ---------- Button Centered with Wesco Green ----------
# Centered, full-width Wesco green button
left, center, right = st.columns([1, 4, 1])
with center:
    run = st.button("🔍 Extract and Compare", use_container_width=True)
    st.markdown("""
        <style>
        div.stButton > button {
            background-color: #00A03E;
            color: white;
            font-size: 1.2em;
            padding: 0.8em;
            border-radius: 8px;
        }
        </style>
    """, unsafe_allow_html=True)

if run:
    debit_col = None
    credit_col = None
    if not image_file or not excel_file:
        st.warning("Please upload both image and Excel files.")
        st.stop()

    # ---------- Process Image ----------
    image_bytes = image_file.read()
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")

    completion = client.chat.completions.create(
        model=deployment_name,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text":
                 "From this journal table image, extract ONLY the final TOTAL Debit and final TOTAL Credit "
                 "values from the summary row at the bottom of the table. Do not extract line items. "
                 "Return just the two numbers."},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}}
            ],
        }],
        max_tokens=200,
    )

    vision_response = completion.choices[0].message.content or ""
    m = re.findall(r"\$?([\d,]+(?:\.\d{2}))", vision_response)
    if len(m) >= 2:
        image_debit = float(m[0].replace(",", ""))
        image_credit = float(m[1].replace(",", ""))
    else:
        image_debit = image_credit = None
        st.error("❌ Could not extract debit/credit from image response.")

    # ---------- Load Excel ----------
    def load_full_excel_as_df(filepath):
        ext = Path(filepath.name).suffix.lower()
        if ext == ".xlsx":
            wb = load_workbook(filepath, data_only=True)
            sheet = wb.active
            data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            return pd.DataFrame(data)
        elif ext == ".xls":
            df_temp = pd.read_excel(filepath, engine="xlrd")
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_temp.to_excel(writer, index=False)
            buffer.seek(0)
            wb = load_workbook(buffer, data_only=True)
            sheet = wb.active
            data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            return pd.DataFrame(data)
        else:
            raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")

    def find_journal_header_row(df_raw):
        for idx, row in df_raw.iterrows():
            row_strs = [str(c).strip().lower() for c in row if pd.notna(c)]
            if any("entered debit" in c for c in row_strs) and any("entered credit" in c for c in row_strs):
                return idx
        return None

    def extract_journal_table(uploaded_file):
        df_raw = load_full_excel_as_df(uploaded_file)
        header_row = find_journal_header_row(df_raw)
        if header_row is None:
            return pd.DataFrame()
        df_journal = df_raw.iloc[header_row:].copy()
        df_journal.columns = df_journal.iloc[0]
        df_journal = df_journal.drop(header_row)
        df_journal = df_journal.reset_index(drop=True)
        df_journal.columns = [str(col).strip() for col in df_journal.columns]
        return df_journal

    df = extract_journal_table(excel_file)

    excel_debit = excel_credit = None
    if not df.empty:
        debit_col = [col for col in df.columns if "entered debit" in col.lower()]
        credit_col = [col for col in df.columns if "entered credit" in col.lower()]

        if debit_col and credit_col:
            debit_col = debit_col[0]
            credit_col = credit_col[0]

            for col in [debit_col, credit_col]:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "").str.replace("$", ""), errors="coerce")

            not_both = ~(df[debit_col].notna() & df[credit_col].notna())
            row_text = df.astype(str).apply(lambda r: " | ".join(r.values).lower(), axis=1)
            not_summary = ~row_text.str.contains(r"\\btotal\\b|\\bsummary\\b", regex=True)

            context_cols = [col for col in df.columns if any(k in col.lower() for k in [
                "legal entity", "operating group", "account", "department", "site",
                "intercompany", "project", "future 1", "future 2", "currency"])]

            required_non_nulls = max(1, len(context_cols) // 2)
            not_mostly_empty = df[context_cols].notna().sum(axis=1) >= required_non_nulls

            filtered = df[not_both & not_summary & not_mostly_empty]
            excel_debit = filtered[debit_col].sum(skipna=True)
            excel_credit = filtered[credit_col].sum(skipna=True)
        else:
            st.error("❌ No backup journal table was found in the uploaded Excel file. Please ensure the Excel contains a table with 'Entered Debit' and 'Entered Credit' columns.")
    else:
        st.error("❌ No backup journal table was found in the uploaded Excel file. Please ensure the Excel contains a table with 'Entered Debit' and 'Entered Credit' columns.")

    # ---------- Comparison Result ----------
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center;'>
        <h2 style='color: #333;'>📌 Comparison Result</h2>
    </div>
    """, unsafe_allow_html=True)

    if None not in (image_debit, image_credit, excel_debit, excel_credit):
        col1, col2 = st.columns(2)
        with col1:
            st.metric(label="🔎 Image Total Debit", value=f"${image_debit:,.2f}")
            st.metric(label="🔎 Image Total Credit", value=f"${image_credit:,.2f}")
        with col2:
            st.metric(label="📊 Excel Total Debit", value=f"${excel_debit:,.2f}")
            st.metric(label="📊 Excel Total Credit", value=f"${excel_credit:,.2f}")

        match = abs(image_debit - excel_debit) < 0.01 and abs(image_credit - excel_credit) < 0.01
        if match:
            st.markdown("""
            <div style='text-align: center; padding: 1em;'>
                <span style='font-size: 1.5em; color: green;'>✅ MATCH: Image and Excel totals are the same.</span>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style='text-align: center; padding: 1em;'>
                <span style='font-size: 1.5em; color: red;'>❌ NO MATCH: The extracted totals do not align.</span>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.warning("⚠️ Unable to determine or compare totals from image and/or Excel.")

    # ---------- Preview Section ----------
    st.markdown("---")
    st.markdown("---")
    st.markdown("---")
    st.markdown("### 🔍 For Reference: Preview Extracted Data")

    # Columns for preview
    prev1, prev2 = st.columns(2)

    with prev1:
        st.markdown("🖼️ **Uploaded Image Preview**", unsafe_allow_html=True)
        # Resize image to fixed height
        img = Image.open(io.BytesIO(image_bytes))
        fixed_height = 400
        w_percent = (fixed_height / float(img.size[1]))
        width = int((float(img.size[0]) * float(w_percent)))
        img = img.resize((width, fixed_height))
        st.image(img, use_container_width=True)

    with prev2:
        st.markdown("📊 **Uploaded Excel Preview**", unsafe_allow_html=True)
        if debit_col and credit_col and not df.empty:
            st.dataframe(df[[debit_col, credit_col]].head(10), use_container_width=True, height=400)
        else:
            st.info("No preview available for Excel debit/credit columns.")
